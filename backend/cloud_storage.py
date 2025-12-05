"""
Data storage module for saving user submissions to Google Cloud Storage.
Works on Cloud Run with persistent storage.
"""

import os
from datetime import datetime
from google.cloud import storage
import openpyxl
from openpyxl import Workbook
import io


def save_user_data(data: dict):
    """
    Save user form data to Excel file in Google Cloud Storage.
    
    Args:
        data: Dictionary containing form data (name, email, topic, etc.)
    """
    try:
        # Get bucket name from environment variable
        bucket_name = os.getenv('GCS_BUCKET_NAME', 'teaching-platform-data')
        file_name = 'user_submissions.xlsx'
        
        print(f"💾 Attempting to save data to GCS bucket: {bucket_name}/{file_name}")
        
        # Initialize GCS client
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(file_name)
        
        # Try to download existing file
        workbook = None
        try:
            excel_data = blob.download_as_bytes()
            workbook = openpyxl.load_workbook(io.BytesIO(excel_data))
            sheet = workbook.active
            print(f"📊 Loaded existing Excel file with {sheet.max_row} rows")
        except Exception as e:
            print(f"📝 Creating new Excel file (existing file not found or error: {e})")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "User Submissions"
            
            # Add headers
            headers = ['Timestamp', 'Name', 'Email', 'Topic', 'Background', 'Experience', 'Goals']
            sheet.append(headers)
        
        # Add new row
        row_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data.get('name', ''),
            data.get('email', ''),
            data.get('topic', ''),
            data.get('background', ''),
            data.get('experience', ''),
            data.get('goals', 'N/A')  # Optional field
        ]
        
        sheet = workbook.active
        sheet.append(row_data)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload to GCS
        blob.upload_from_file(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        print(f"✅ Successfully saved user data to GCS: {bucket_name}/{file_name}")
        print(f"   Total rows: {sheet.max_row}")
        
    except Exception as e:
        print(f"❌ Error saving to GCS: {type(e).__name__}: {str(e)}")
        # Don't raise - we don't want to fail the request if storage fails
        print("⚠️  Continuing without saving to Excel (email was still sent)")
