from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Get absolute path to data directory
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'user_submissions.xlsx')

def save_user_data(data):
    """Save user form submission to Excel file."""
    
    try:
        # Create data directory if it doesn't exist
        os.makedirs(DATA_DIR, exist_ok=True)
        print(f"📁 Data directory: {DATA_DIR}")
        print(f"📄 Excel file path: {EXCEL_FILE}")
        
        # Check if file exists
        if os.path.exists(EXCEL_FILE):
            print(f"📂 Loading existing Excel file...")
            # Load existing workbook
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            print(f"📝 Creating new Excel file...")
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "User Submissions"
            
            # Add headers
            headers = [
                'Timestamp',
                'Name',
                'Email',
                'Topic',
                'Background',
                'Experience Level',
                'Specific Interest',
                'Learning Goal'
            ]
            ws.append(headers)
        
        # Add new row with user data
        row_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data.get('name', ''),
            data.get('email', ''),
            data.get('topic', ''),
            data.get('background', ''),
            data.get('experience', ''),
            data.get('interest', ''),
            data.get('goal', '')
        ]
        
        ws.append(row_data)
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        print(f"✅ User data saved to Excel: {data.get('email')}")
        print(f"📊 Total rows in Excel: {ws.max_row}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
